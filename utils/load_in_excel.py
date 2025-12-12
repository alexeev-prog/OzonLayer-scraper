# utils/load_in_excel.py
import logging

import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from utils.config import CONFIG


def write_data_to_excel(
    products_data: dict[str, dict[str, str | None | dict | list]],
    filename: str | None = None,
) -> None:
    if not products_data:
        logging.warning("Передан пустой словарь данных. Операция записи пропущена.")
        return

    excel_config = CONFIG["excel"]
    filename = filename or excel_config["filename"]
    sheet_name = excel_config["sheet_name"]

    try:
        all_char_keys = set()
        for product_info in products_data.values():
            characteristics = product_info.get("product_characteristics", {})
            all_char_keys.update(characteristics.keys())
        sorted_char_keys = sorted(all_char_keys)

        flattened_data = []
        for product_id, product_info in products_data.items():
            flat_product = {
                "product_id": product_id,
                "product_name": product_info.get("product_name"),
                "product_ozon_card_price": product_info.get("product_ozon_card_price"),
                "product_discount_price": product_info.get("product_discount_price"),
                "product_base_price": product_info.get("product_base_price"),
                "product_stars": product_info.get("product_stars"),
                "product_reviews": product_info.get("product_reviews"),
                "salesman": product_info.get("salesman"),
                "product_url": product_info.get("product_url"),
            }

            characteristics = product_info.get("product_characteristics", {})
            for key in sorted_char_keys:
                flat_product[f"Характеристика {key}"] = characteristics.get(key)

            colors = product_info.get("product_colors", [])
            colors = [color for color in colors if color and color.strip()]
            flat_product["product_colors"] = ", ".join(colors) if colors else None

            flattened_data.append(flat_product)

        df = pd.DataFrame(flattened_data)

        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            worksheet = writer.sheets[sheet_name]

            for col_idx, column_cells in enumerate(worksheet.iter_cols(), start=1):
                max_length = max(
                    (len(str(cell.value)) for cell in column_cells if cell.value),
                    default=0,
                )
                worksheet.column_dimensions[get_column_letter(col_idx)].width = (
                    max_length + 2
                )

            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

        logging.info(f"Файл '{filename}' успешно создан!")
    except Exception as e:
        logging.exception(f"Ошибка при записи в файл '{filename}': {e}")
        raise
